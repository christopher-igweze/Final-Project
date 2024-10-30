import glob
import openpyxl
import sqlite3
import itertools
# from Frontend.pages.createTT import alpha
import os

alpha = False

master_folder = '/config/Final-Project/Version 3.0/extracted_folder'

# Get all folders in the directory
folders = [folder for folder in os.listdir(master_folder) if os.path.isdir(os.path.join(master_folder, folder))]

folder_path = os.path.join(master_folder, folders[0])

db_file = '/config/Final-Project/Version 3.0/class_schedule-02.db'

# Create a connection with the db
conn = sqlite3.connect(db_file)
cursor = conn.cursor()
# Clear the 'course' table
cursor.execute("DELETE FROM course")
cursor.execute("DELETE FROM course_student")
cursor.execute("DELETE FROM course_instructor")
counter = 1
Scounter = 1

# Fetch the data from the 'instructor' table
cursor.execute("SELECT * FROM instructor")
data = cursor.fetchall()

# Create a dictionary to store the data
instructor_dict = {}
for row in data:
    instructor_dict[row[1]] = row[0]

course_dict = {}
students = {}
student_idList = []
# Loop through the files in the master folder
for file_path in glob.glob(folder_path + '/*.xlsx'):
    # Loop through the sheets of each workbook
    workbook = openpyxl.load_workbook(file_path)
    if alpha == False:
        actual = [sheet_name for sheet_name in workbook.sheetnames if sheet_name not in ['300', '500']]
        for sheet_name in actual:
            sheet = workbook[sheet_name]      
            # Populate the table
            for i in range(2, sheet.max_row + 1):
                course_code = f"C{counter}"
                counter += 1
                name = sheet.cell(row=i, column=1).value
                if not name:
                    continue
                name = str(name)
                numberOfStudents = sheet.cell(row=i, column=2).value
                units = sheet.cell(row=i, column=3).value
                instructors = str(sheet.cell(row=i, column=4).value).split(',')
                elective = sheet.cell(row=i, column=6).value
                course_dict[name] = course_code
                
                # Check if the course already exists in the database
                cursor.execute("SELECT * FROM course WHERE name=?", (name,))
                existing_course = cursor.fetchone()
                
                if existing_course:
                    # Course already exists, update the total number of students
                    total_students = existing_course[2] + numberOfStudents
                    cursor.execute("UPDATE course SET max_numb_of_students=? WHERE name=?", (total_students, name))
                else:
                    # Course does not exist, insert a new record
                    cursor.execute("SELECT * FROM dept")
                    thisData = cursor.fetchall()
                    dept_row = next((_ for _ in thisData if _[2] == sheet.cell(row=i, column=5).value), None)
                    if dept_row:
                        # populate instructor table
                        cursor.execute("SELECT * FROM instructor")
                        insData = cursor.fetchall()
                        prev_id = int(insData[-1][0][1:])
                        dept_id = dept_row[2]
                        new_instructors = []
                        for instructor in instructors:
                            if instructor not in instructor_dict:
                                new_id = f"I{prev_id + 1}"
                                new_instructors.append((new_id, instructor, dept_id))
                                instructor_dict[instructor] = new_id
                                prev_id += 1
                        
                        cursor.executemany("INSERT INTO instructor (number, name, Dept_id) VALUES (?, ?, ?)", new_instructors)

                        dept_value = dept_row[0]
                        cursor.execute("INSERT INTO course (number, name, max_numb_of_students, credit_hours, Elective) VALUES (?, ?, ?, ?, ?)", (course_code, name, numberOfStudents, units, elective))
                        cursor.execute("INSERT INTO dept_course (name, course_numb) VALUES (?, ?)", (dept_value, course_code))
                        for instructor in instructors:
                            cursor.execute("INSERT INTO course_instructor (course_number, instructor_number) VALUES (?, ?)", (course_code, instructor_dict.get(instructor)))
                    else:
                        # Handle the case where the department is not found
                        print(f"Department not found for the course: {name}")

            
            # Get the number of electives required for the semester
            noOfElectives = sheet.cell(row=2, column=7).value
            compulsory_list = [sheet.cell(row=i, column=1).value for i in range(2, sheet.max_row + 1) if sheet.cell(row=i, column=6).value == "NO"]
            elective_list = [sheet.cell(row=i, column=1).value for i in range(2, sheet.max_row + 1) if sheet.cell(row=i, column=6).value == "YES"]
            
            # Ensure noOfElectives is an integer and not None before using it
            noOfElectives = int(noOfElectives) if noOfElectives is not None else 0 # type: ignore

            for combination in itertools.combinations(elective_list, min(len(elective_list), noOfElectives)):
                student_code = f"S{Scounter}"
                Scounter += 1
                student_list = compulsory_list + list(combination)
                student_idList = [course_dict.get(course) for course in student_list if course_dict.get(course) is not None]
                students[student_code] = student_idList
            print(students)

        for id, theList in students.items():
            for course_id in theList:
                cursor.execute("INSERT INTO course_student (course_id, student_id) VALUES (?, ?)", (course_id, id))
          
conn.commit()
cursor.close()
# Close the connection
conn.close()

